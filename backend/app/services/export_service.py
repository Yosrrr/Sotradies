"""Génération des exports Excel et PDF — marchés et journal d'audit."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

HEADER_FILL = PatternFill(start_color="1E2F49", end_color="1E2F49", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Caractères qu'Excel/LibreOffice peuvent interpréter comme le début
# d'une formule si un contenu externe non fiable (scraping) les contient
# en première position (OWASP "CSV/Formula Injection").
FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _sanitize_excel_cell(value):
    """Neutralise une valeur texte qui pourrait être interprétée comme
    une formule Excel — préfixe d'une apostrophe, sans changer les
    valeurs numériques (score, etc.) qui n'ont pas ce risque."""
    if isinstance(value, str) and value.startswith(FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def _truncate(text: str | None, length: int) -> str:
    if not text:
        return "-"
    return text if len(text) <= length else text[: length - 1] + "…"


# ---------- MARCHÉS ----------

def tenders_to_excel(tenders: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Marchés"

    headers = ["ID", "Objet", "Acheteur", "Catégorie", "Score (%)", "Statut",
           "Commercial", "Source", "Date publication", "Date limite", "Lien"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for t in tenders:
        ws.append([
            _sanitize_excel_cell(t.id or ""),
             _sanitize_excel_cell(t.objet or ""),
        _sanitize_excel_cell(t.acheteur or ""),
        _sanitize_excel_cell(t.categorie or ""),
        t.score,
        t.statut,
        _sanitize_excel_cell(t.commercial_assigne or ""),
        _sanitize_excel_cell(t.source or ""),
        t.date_publication.strftime("%Y-%m-%d") if t.date_publication else "",
        t.date_limite.strftime("%Y-%m-%d") if t.date_limite else "",
        _sanitize_excel_cell(t.lien or ""),
        ])

    widths = [16, 45, 30, 20, 10, 12, 18, 16, 16, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def tenders_to_pdf(tenders: list) -> bytes:
    # Pas de sanitation nécessaire ici : un PDF n'exécute jamais de
    # formule, le risque est spécifique aux tableurs (Excel/LibreOffice).
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Marchés — SOTRADIES")
    styles = getSampleStyleSheet()
    elements = [Paragraph("Marchés — Veille des appels d'offres SOTRADIES", styles["Title"]), Spacer(1, 10)]

    data = [["Objet", "Acheteur", "Catégorie", "Score", "Statut", "Commercial", "Date limite"]]
    for t in tenders:
        data.append([
            _truncate(t.objet, 55),
            _truncate(t.acheteur, 28),
            t.categorie or "-",
            f"{t.score}%",
            t.statut,
            t.commercial_assigne or "-",
            t.date_limite.strftime("%d/%m/%Y") if t.date_limite else "-",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E2F49")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F7F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


# ---------- JOURNAL D'AUDIT ----------

def audit_log_to_excel(logs: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal d'audit"

    headers = ["Date", "Utilisateur", "Action", "Marché concerné", "Détail"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for log in logs:
        ws.append([
            log.date_action.strftime("%Y-%m-%d %H:%M:%S"),
            _sanitize_excel_cell(log.utilisateur_email),
            log.action,
            _sanitize_excel_cell(log.tender_objet or ""),
            _sanitize_excel_cell(log.detail or ""),
        ])

    widths = [20, 28, 20, 45, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def audit_log_to_pdf(logs: list) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Journal d'audit — SOTRADIES")
    styles = getSampleStyleSheet()
    elements = [Paragraph("Journal d'audit — Veille des appels d'offres SOTRADIES", styles["Title"]), Spacer(1, 10)]

    data = [["Date", "Utilisateur", "Action", "Marché concerné"]]
    for log in logs:
        data.append([
            log.date_action.strftime("%d/%m/%Y %H:%M"),
            log.utilisateur_email,
            log.action,
            _truncate(log.tender_objet or log.detail, 50),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E2F49")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F7F9")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()