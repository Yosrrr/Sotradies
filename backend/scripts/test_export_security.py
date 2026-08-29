from io import BytesIO
from types import SimpleNamespace
from datetime import datetime
from openpyxl import load_workbook
from app.services.export_service import tenders_to_excel, audit_log_to_excel


def test_tender_excel_formula_injection_is_neutralized():
    t = SimpleNamespace(
        id="1", objet='=HYPERLINK("http://evil")', acheteur="+MAL",
        categorie="-BAD", score=90, statut="retenu",
        commercial_assigne="@ATK", source="onmp",
        date_publication=datetime(2026, 8, 21),
        date_limite=datetime(2026, 8, 30), lien="http://example.com",
    )
    wb = load_workbook(BytesIO(tenders_to_excel([t])), data_only=False)
    ws = wb["Marchés"]
    assert ws["B2"].value.startswith("'=")
    assert ws["C2"].value.startswith("'+")
    assert ws["D2"].value.startswith("'-")
    assert ws["G2"].value.startswith("'@")


def test_audit_excel_formula_injection_is_neutralized():
    log = SimpleNamespace(
        date_action=datetime(2026, 8, 21, 10),
        utilisateur_email="=evil@x.com", action="consultation",
        tender_objet="+bad", detail="@bad",
    )
    wb = load_workbook(BytesIO(audit_log_to_excel([log])), data_only=False)
    ws = wb["Journal d'audit"]
    assert ws["B2"].value.startswith("'=")
    assert ws["D2"].value.startswith("'+")
    assert ws["E2"].value.startswith("'@")